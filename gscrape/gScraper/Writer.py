from openpyxl import load_workbook

# wb = load_workbook("input1.xlsx")
# 
# # grab the active worksheet
# ws = wb.get_sheet_by_name('Feuil1')



# for row in ws.rows:
#     row[0].value = 12

#wb.save('input1.xlsx')

def writeFile(inputFile, wordList):
    wb = load_workbook(inputFile)
    # grab the active worksheet
    ws = wb.get_sheet_by_name('MainDatafile')
    
    i = 0
    for row in ws.rows[1:]:
        try:
            row[19].value = wordList[i]['beer']
            row[20].value = wordList[i]['wine']
            row[21].value = wordList[i]['alcohol']
        except:
            pass
        i = i+1
    
    print 'Writing file'
    while True:
        try:
            wb.save(inputFile)
            break
        except:
            pass
        
    print 'Writing successfull'